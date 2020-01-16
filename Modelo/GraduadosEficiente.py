import pandas as pd
from openpyxl import load_workbook
import pandas.io.formats.excel
import xlrd

#Funcion para agregar columna
def agrega_Columna(nombre,nombre_hoja,valores_agregar,numero_columna):
    book = load_workbook(nombre)
    df = pd.DataFrame(valores_agregar)
    columnas = []
    for i in valores_agregar:
        columnas.append(i)
    df = df[columnas]
    writer = pd.ExcelWriter(nombre)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=nombre_hoja, startcol=numero_columna, index=False)
    writer.save()

#Variables a cargar en memoria
ano_ingreso = []
fecha_egreso = []
termino_ingreso = []
termino_egreso = []

# Configuramos Pandas y cargamos el archivo correspondiente (en este caso se llama archivo.xlsx)   
nombre = "..\\Excel_generados\\graduados.xlsx"
openfile = xlrd.open_workbook(nombre)
hoja = openfile.sheet_by_name("estudiantes_graduados")
for i in range(hoja.nrows):
    if i!= 0:
        aingreso = int((hoja.cell_value(i,4)))
        fegreso = int((hoja.cell_value(i,5)))
        tingreso = str((hoja.cell_value(i,6))).split(" ")[0]
        tegreso = str((hoja.cell_value(i,7))).split(" ")[0]
        ano_ingreso.append(aingreso)
        fecha_egreso.append(fegreso)
        termino_ingreso.append(tingreso)
        termino_egreso.append(tegreso)
def definir_Eficiencia(aingreso,fegreso,tingreso,tegreso,tiempo):
    eficiencia = []
    for i in range(len(aingreso)):
        diferencia = fegreso[i] - aingreso[i]
        t_egreso = tegreso[i]
        t_ingreso = tingreso[i]
        if((diferencia == tiempo+1 and t_ingreso.lower()=='2s' and t_egreso.lower()=='1s') or diferencia<=tiempo):
            eficiencia.append(1)
        else:
            eficiencia.append(0)
    return eficiencia

eficiencia = definir_Eficiencia(ano_ingreso,fecha_egreso,termino_ingreso,termino_egreso,5)

diccionario = {"Eficiencia_Terminal": eficiencia}

agrega_Columna(nombre,"estudiantes_graduados",diccionario,8)





'''# Configuramos Pandas y cargamos el archivo correspondiente (en este caso se llama archivo.xlsx)                       
book = load_workbook('archivo.xlsx')
writer = pd.ExcelWriter('archivo.xlsx', engine='openpyxl') 
writer.book = book

# Por defecto Pandas formatea las celdas del header con negrita y borde, si no se quiere hacemos lo siguiente:
pandas.io.formats.excel.header_style = None

# Guardamos el df en el excel en el lugar apropiado.
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
df.to_excel(writer, book.worksheets[0].title, startcol = 5,  index = False) 
writer.save()




def agregarColumnaExcel(nombre,hoja,columna,valores){

}'''